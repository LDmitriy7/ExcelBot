from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import date, timedelta


def get_date_from_term(term: str) -> date:
    today = date.today()
    if term == 'на сегодня':
        return today
    weeks = int(term.split()[1])
    new_date = today + timedelta(weeks=weeks)
    days_from_Monday = new_date.weekday()
    new_date = new_date - timedelta(days=days_from_Monday)
    return new_date


class ExcelBook:
    def __init__(self, wb_name):
        self.wb_name = wb_name
        self.unlisted_name = 'Учет невыложенных'
        self.in_reserve_name = 'Учет резерва'

    def get_wb(self) -> Workbook:
        wb = load_workbook(self.wb_name)
        return wb

    def add_to_unlisted(self, row):
        wb = self.get_wb()
        unlisted_ws = wb[self.unlisted_name]
        unlisted_ws.append(row)
        wb.save(self.wb_name)

    def add_to_reserve(self, row):
        wb = self.get_wb()
        in_reserve_ws = wb[self.in_reserve_name]
        in_reserve_ws.append(row)
        wb.save(self.wb_name)

    def get_reserve(self):
        wb = self.get_wb()
        in_reserve_ws = wb[self.in_reserve_name]
        rows = in_reserve_ws.iter_rows(2, min_col=3, max_col=6, values_only=True)
        records = []
        for record in rows:
            if record[0] is None:
                break
            records.append(record)
        cols = zip(*records)
        results = [int(sum(i)) for i in cols]
        return results

    def get_unlisted(self, term: str):
        """term может быть = [на сегодня, на 1 неделю, на 2 недели, на 3 недели, на 4 недели]"""
        wb = self.get_wb()
        unlisted_ws: Worksheet = wb[self.unlisted_name]
        term = get_date_from_term(term)
        rows = unlisted_ws.iter_rows(2, min_col=4, max_col=8, values_only=True)
        fit_records = []
        for record in rows:
            if record[0] is None:
                break
            amounts = list(record)
            record_date = amounts.pop()
            if record_date is None:
                fit_records.append(amounts)
            elif record_date.date() <= term:
                fit_records.append(amounts)

        cols = zip(*fit_records)
        results = [int(sum(i)) for i in cols]
        return results
